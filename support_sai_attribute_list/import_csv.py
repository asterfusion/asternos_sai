import csv
import re
import os
from openpyxl import Workbook

def md_table_to_csv(md_content):
    lines = md_content.strip().split('\n')
    headers = re.split(r'\s*\|\s*', lines[0].strip())
    headers = [h.strip() for h in headers]  # Preserve empty cells

    data = []
    for line in lines[2:]:
        if re.match(r'^\s*[-|]+\s*$', line):
            continue
        row = re.split(r'\s*\|\s*', line.strip())
        row = [cell.strip() for cell in row]  # Preserve empty cells
        data.append(row)

    return headers, data

def write_to_csv(headers, data, csv_file):
    # Remove the first column from headers and data
    headers = headers[1:] if headers else []
    data = [row[1:] for row in data]

    with open(csv_file, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(data)

def convert_md_to_csv(md_file, csv_file):
    with open(md_file, 'r', encoding='utf-8') as f:
        md_content = f.read()
    headers, data = md_table_to_csv(md_content)
    write_to_csv(headers, data, csv_file)

def merge_csv_files(output_file):
    wb = Workbook()
    # Remove the default sheet created by Workbook()
    if "Sheet" in wb.sheetnames:
        default_sheet = wb["Sheet"]
        wb.remove(default_sheet)

    # Get all .csv files in the current directory
    for root, dirs, files in os.walk('.'):
        for file in files:
            if file.endswith('.csv'):
                csv_file = os.path.join(root, file)
                # Use the file name without extension as sheet name
                sheet_name = os.path.splitext(os.path.basename(csv_file))[0]
                # Truncate sheet name to 31 characters as Excel limit
                sheet_name = sheet_name[:31]
                sheet = wb.create_sheet(sheet_name)

                with open(csv_file, 'r', encoding='utf-8') as f:
                    reader = csv.reader(f)
                    for row in reader:
                        sheet.append(row)

    # Save the workbook
    wb.save(output_file)

if __name__ == "__main__":
    # Convert all .md files to .csv files
    for root, dirs, files in os.walk('.'):
        for file in files:
            if file.endswith('.md'):
                md_file = os.path.join(root, file)
                csv_file = os.path.splitext(md_file)[0] + '.csv'
                convert_md_to_csv(md_file, csv_file)

    # Merge all .csv files into one Excel file
    merge_csv_files('merged.xlsx')